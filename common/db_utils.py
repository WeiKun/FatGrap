# -*- coding: utf-8 -*-
import os
import threading
import openpyxl
import utils
import global_var
import traceback


class DBImp(object):
    @staticmethod
    def read_column(row_key, column_key, default=''):
        pass

    @staticmethod
    def write_row(row_key, column_key, value):
        pass


class XlsDB(DBImp):
    fileName = ''
    wb = None
    ws = None
    major_key = ''
    locker = None
    column_dict = {}

    @staticmethod
    def format(string):
        if string:
            if type(string) == str:
                return string.decode(global_var.LOCAL_CODE)
            elif type(string) == unicode:
                return string
            else:
                return str(string).decode(global_var.LOCAL_CODE)
        else:
            return string

    @staticmethod
    def reformat(string):
        if string:
            if type(string) == str:
                return string.encode(global_var.LOCAL_CODE)
            else:
                return str(string)
        else:
            return string

    @staticmethod
    def init():
        if XlsDB.fileName or XlsDB.wb:
            return

        XlsDB.fileName = utils.join_file_path(global_var.XLSX_NAME)

        if not os.path.exists(XlsDB.fileName):
            XlsDB.wb = openpyxl.Workbook(XlsDB.fileName)
            XlsDB.save()
            XlsDB.wb = openpyxl.load_workbook(XlsDB.fileName)
        else:
            XlsDB.wb = openpyxl.load_workbook(XlsDB.fileName)

        XlsDB.ws = XlsDB.wb.active
        XlsDB.ws.title = "main"
        for i, column_name in enumerate(global_var.XLSX_COLUMNS):
            column = i + 1
            column_name = XlsDB.format(column_name)
            XlsDB.ws.cell(1, column).value = column_name
            XlsDB.column_dict[column_name] = column
        XlsDB.major_key = global_var.XLSX_COLUMNS[0]
        XlsDB.save()

        XlsDB.locker = threading.Lock()

    @staticmethod
    def save():
        XlsDB.wb.save(XlsDB.fileName)

    @staticmethod
    def read_column(row_key, column_key, default=''):
        XlsDB.locker.acquire()
        try:
            ret = XlsDB._read_column(row_key, column_key, default)
        except Exception as exception:
            traceback.print_exc()
            XlsDB.locker.release()
            raise exception
        XlsDB.locker.release()
        return ret

    @staticmethod
    def _read_column(row_key, column_key, default=''):
        row_key = XlsDB.format(row_key)
        column_key = XlsDB.format(column_key)

        max_row = XlsDB.ws.max_row
        for row_index in xrange(2, max_row + 1):
            majoy_cell = XlsDB.ws.cell(row_index, 1)
            if majoy_cell.value == row_key:
                column_index = XlsDB.column_dict[column_key]
                return XlsDB.reformat(XlsDB.ws.cell(row_index, column_index).value)

        return XlsDB.reformat(default)

    @staticmethod
    def write_row(row_key, **kw):
        XlsDB.locker.acquire()
        try:
            XlsDB._write_row(row_key, **kw)
        except Exception as exception:
            traceback.print_exc()
            XlsDB.locker.release()
            raise exception
        XlsDB.locker.release()

    @staticmethod
    def _write_row(row_key, **kw):
        row_key = str(row_key)

        max_row = XlsDB.ws.max_row
        for row_index in xrange(2, max_row + 1):
            majoy_cell = XlsDB.ws.cell(row_index, 1)
            if majoy_cell.value == XlsDB.format(row_key):
                for column_key, value in kw.iteritems():
                    column_index = XlsDB.column_dict[XlsDB.format(column_key)]
                    XlsDB.ws.cell(
                        row_index, column_index).value = XlsDB.format(value)
                break
        else:
            row_index = max_row + 1
            XlsDB.ws.cell(row_index, 1).value = XlsDB.format(row_key)
            for column_key, value in kw.iteritems():
                column_index = XlsDB.column_dict[XlsDB.format(column_key)]
                XlsDB.ws.cell(
                    row_index, column_index).value = XlsDB.format(value)
        XlsDB.save()
